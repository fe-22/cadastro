import openpyxl

class GeradorExcel:
    def __init__(self, arquivo_nome):
        self.arquivo_nome = arquivo_nome
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active

    def escrever_dados(self, dados):
        for dado in dados:
            self.sheet.append(dado)

    def salvar(self):
        self.workbook.save(self.arquivo_nome)

# Criar uma planilha de exemplo
gerador_excel = GeradorExcel('example.xlsx')
cabecalho = ['Nome', 'CPF', 'Data de Nascimento', 'Sexo', 'Estado Civil', 'Profissão', 'Escolaridade', 
             'Data de Batismo', 'Data de Conversão', 'Data de Ingresso', 'Endereço Residencial', 
             'Telefone', 'Email', 'Cargo na Igreja', 'Área de Ministério', 'Nome do Cônjuge', 'Filhos',
               'Alergias', 'Doenças Crônicas', 'Observações']
gerador_excel.escrever_dados([cabecalho])  # Adicionando colchetes extras para passar a lista como um único elemento
gerador_excel.salvar()

def ler_dados_excel(arquivo_nome):
    workbook = openpyxl.load_workbook(arquivo_nome)
    sheet = workbook.active

    dados = []
    for row in sheet.iter_rows(values_only=True):
        dados.append(row)

    return dados

# Ler os dados da planilha
dados_excel = ler_dados_excel('example.xlsx')
print(dados_excel)

