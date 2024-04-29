import tkinter as tk
from tkinter import ttk
import os
import openpyxl
from tkinter import messagebox
import matplotlib.pyplot as plt

from collections import Counter

from ger_graf import GeradorGrafico
from ger_ex import GeradorExcel


# Classe GeradorExcel
class GeradorExcel:
    def __init__(self, arquivo_nome, diretorio=None):
        self.arquivo_nome = arquivo_nome
        self.diretorio = diretorio

    def escrever_dados(self, dados):
        caminho_completo = os.path.join(self.diretorio, self.arquivo_nome)
        if not os.path.exists(caminho_completo):
            workbook = openpyxl.Workbook()
            workbook.save(caminho_completo)
        workbook = openpyxl.load_workbook(caminho_completo)
        sheet = workbook.active
        for linha in dados:
            sheet.append(linha)
        workbook.save(caminho_completo)

    def definir_diretorio(self, diretorio):
        self.diretorio = diretorio

    def salvar(self):
        if self.diretorio:
            caminho_completo = os.path.join(self.diretorio, self.arquivo_nome)
            caminho_completo_com_extensao = caminho_completo + ".xlsx"
            
            if os.path.exists(caminho_completo_com_extensao):
                workbook = openpyxl.load_workbook(caminho_completo_com_extensao)
                for sheet in workbook.sheetnames:
                    workbook[sheet].delete_rows(1, workbook[sheet].max_row)
                workbook.save(caminho_completo_com_extensao)
            else:
                workbook = openpyxl.Workbook()
                workbook.save(caminho_completo_com_extensao)
        else:
            raise ValueError("O diretório de salvamento não foi especificado.")


# Função relacionada ao cadastro de membros
def cadastrar_novo_membro():
    try:
        # Obter os valores dos campos
        nome = nome_var.get()
        cpf = cpf_var.get()
        data_nascimento = data_nascimento_var.get()
        sexo = sexo_var.get()
        estado_civil = estado_civil_var.get()
        profissao = profissao_var.get()
        escolaridade = escolaridade_var.get()
        data_batismo = data_batismo_var.get()
        data_conversao = data_conversao_var.get()
        data_ingresso = data_ingresso_var.get()
        endereco_residencial = endereco_residencial_var.get()
        telefone = telefone_var.get()
        email = email_var.get()
        # Campos adicionais
        cargo_na_igreja = cargo_na_igreja_var.get()
        area_ministerio = area_ministerio_var.get()
        nome_conjuge = nome_conjuge_var.get()
        filhos = filhos_var.get()
        alergias = alergias_var.get()
        doencas_cronicas = doencas_cronicas_var.get()
        observacoes = observacoes_var.get()

        # Salvar os dados no Excel
        gerador_excel.escrever_dados([[nome, cpf, data_nascimento, sexo, estado_civil, profissao, escolaridade,
                                       data_batismo, data_conversao, data_ingresso, endereco_residencial,
                                       telefone, email, cargo_na_igreja, area_ministerio, nome_conjuge,
                                       filhos, alergias, doencas_cronicas, observacoes]])

        # Salvar o arquivo Excel
        gerador_excel.salvar()

        messagebox.showinfo("Sucesso", "Novo membro cadastrado com sucesso!")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao cadastrar o novo membro: {e}")


# Funções relacionadas à exibição de gráficos

def ler_dados_excel(example_nome):
    workbook = openpyxl.load_workbook(example_nome, read_only=True)
    sheet = workbook.active
    dados = [row for row in sheet.iter_rows(values_only=True)]
    cabecalho = dados[0]
    dados_dict = {cabecalho[i]: [row[i] for row in dados[1:]] for i in range(len(cabecalho))}
    return dados_dict


def exibir_graficos():
    try:
        # Ler os dados do arquivo Excel
        dados = ler_dados_excel("example.xlsx")
        
        # Criar uma instância da classe GeradorGrafico
        gerador_grafico = GeradorGrafico(dados)
        
        # Criar e mostrar o gráfico de distribuição por sexo
        gerador_grafico.criar_grafico_pizza_sexo()
        
        # Criar e mostrar o gráfico de distribuição por estado civil
        gerador_grafico.criar_grafico_pizza_estado_civil()
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao exibir os gráficos: {e}")
def exibir_graficos():
    try:
        # Ler os dados do arquivo Excel
        dados = ler_dados_excel("example.xlsx")
        
        # Criar uma instância da classe GeradorGrafico
        gerador_grafico = GeradorGrafico(dados)
        
        # Criar e mostrar o gráfico de distribuição por sexo
        gerador_grafico.criar_grafico_pizza_sexo()
        
        # Criar e mostrar o gráfico de distribuição por estado civil
        gerador_grafico.criar_grafico_pizza_estado_civil()
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao exibir os gráficos: {e}")
def criar_grafico_pizza_sexo(self):
    sexo_index = self.cabecalho.index("Sexo")
    sexo_data = [row[sexo_index] for row in self.dados]
    sexo_count = Counter(sexo_data)
    plt.figure(figsize=(6, 6))
    plt.pie(sexo_count.values(), labels=sexo_count.keys(), autopct='%1.1f%%', startangle=140)
    plt.title("Distribuição por Sexo")
    plt.axis('equal')
    plt.show()

def criar_grafico_pizza_estado_civil(self):
    estado_civil_index = self.cabecalho.index("Estado Civil")
    estado_civil_data = [row[estado_civil_index] for row in self.dados]
    estado_civil_count = Counter(estado_civil_data)
    plt.figure(figsize=(6, 6))
    plt.pie(estado_civil_count.values(), labels=estado_civil_count.keys(), autopct='%1.1f%%', startangle=140)
    plt.title("Distribuição por Estado Civil")
    plt.axis('equal')
    plt.show()



# Configuração da interface gráfica
app = tk.Tk()
app.title("Cadastro de Membro")

# Cabeçalho 1
header1 = tk.Label(app, text="Cadastro de Membro", font=("Arial", 14, "bold"))
header1.grid(row=0, column=0, columnspan=2, pady=10)

# Cabeçalho 2
header2 = tk.Label(app, text="ASSEMBLEIA DE DEUS - MINISTÉRIO FIDELIDADE", font=("Arial", 14, "bold"))
header2.grid(row=1, column=0, columnspan=2, pady=10)

# Variáveis para armazenar os valores dos campos
nome_var = tk.StringVar()
cpf_var = tk.StringVar()
data_nascimento_var = tk.StringVar()
sexo_var = tk.StringVar()
estado_civil_var = tk.StringVar()
profissao_var = tk.StringVar()
escolaridade_var = tk.StringVar()
data_batismo_var = tk.StringVar()
data_conversao_var = tk.StringVar()
data_ingresso_var = tk.StringVar()
endereco_residencial_var = tk.StringVar()
telefone_var = tk.StringVar()
email_var = tk.StringVar()

# Campos adicionais
cargo_na_igreja_var = tk.StringVar()
area_ministerio_var = tk.StringVar()
nome_conjuge_var = tk.StringVar()
filhos_var = tk.StringVar()
alergias_var = tk.StringVar()
doencas_cronicas_var = tk.StringVar()
observacoes_var = tk.StringVar()

# Frame para os campos do formulário
form_frame = tk.Frame(app)
form_frame.grid(row=2, column=0, columnspan=2, padx=10, pady=10)

# Campos Dados Básicos
label_nome = tk.Label(form_frame, text="Nome Completo:")
label_nome.grid(row=1, column=0, sticky="e", padx=(0, 10))
entry_nome = tk.Entry(form_frame, textvariable=nome_var, width=100)
entry_nome.grid(row=1, column=1)

label_cpf = tk.Label(form_frame, text="CPF:")
label_cpf.grid(row=2, column=0, sticky="e", padx=(10, 10))
entry_cpf = tk.Entry(form_frame, textvariable=cpf_var, width=80)
entry_cpf.grid(row=2, column=1)

label_data_nascimento = tk.Label(form_frame, text="Data de Nascimento:")
label_data_nascimento.grid(row=3, column=0, sticky="e", padx=(0,  10))
entry_data_nascimento = tk.Entry(form_frame, textvariable=data_nascimento_var, width=50)
entry_data_nascimento.grid(row=3, column=1)

label_endereco_residencial = tk.Label(form_frame, text="Endereço Residencial:")
label_endereco_residencial.grid(row=4, column=0, sticky="e", padx=(0, 10))
entry_endereco_residencial = tk.Entry(form_frame, textvariable=endereco_residencial_var, width=100)
entry_endereco_residencial.grid(row=4, column=1)

label_telefone = tk.Label(form_frame, text="Telefone:")
label_telefone.grid(row=5, column=0, sticky="e", padx=(100, 10))
entry_telefone = tk.Entry(form_frame, textvariable=telefone_var, width=50)
entry_telefone.grid(row=5, column=1)

label_email = tk.Label(form_frame, text="Email:")
label_email.grid(row=6, column=0, sticky="e", padx=(0, 10))
entry_email = tk.Entry(form_frame, textvariable=email_var, width=100)
entry_email.grid(row=6, column=1)

# Variável para armazenar o sexo selecionado
sexo_var = tk.StringVar()
sexo_var.set("Masculino")  # Valor padrão inicial

label_sexo = tk.Label(form_frame, text="Sexo:")
label_sexo.grid(row=7, column=0, sticky="e", padx=(0, 10))

# Botões de rádio
radio_masculino = tk.Radiobutton(form_frame, text="Masculino", variable=sexo_var, value="Masculino")
radio_masculino.grid(row=7, column=1, sticky="w")

radio_feminino = tk.Radiobutton(form_frame, text="Feminino", variable=sexo_var, value="Feminino")
radio_feminino.grid(row=8, column=1, sticky="w")

# Combobox para o estado civil
label_estado_civil = tk.Label(form_frame, text="Estado Civil:")
label_estado_civil.grid(row=9, column=0, sticky="e", padx=(0, 10))

# Opções para o combobox
opcoes_estado_civil = ["Casado", "Solteiro", "Viúvo", "Divorciado"]

# Combobox
combo_estado_civil = ttk.Combobox(form_frame, values=opcoes_estado_civil, textvariable=estado_civil_var, width=47)
combo_estado_civil.current(0)  # Definir o valor padrão inicial
combo_estado_civil.grid(row=9, column=1, sticky="w")

label_conjuge = tk.Label(form_frame, text="Nome do cônjuge:")
label_conjuge.grid(row=10, column=0, sticky="e", padx=(0, 10))
entry_conjuge = tk.Entry(form_frame, textvariable=nome_conjuge_var, width=100)
entry_conjuge.grid(row=10, column=1)

label_profissao = tk.Label(form_frame, text="Profissão:")
label_profissao.grid(row=11, column=0, sticky="e", padx=(0, 10))
entry_profissao = tk.Entry(form_frame, textvariable=profissao_var, width=100)
entry_profissao.grid(row=11, column=1)

label_escolaridade = tk.Label(form_frame, text="Escolaridade:")
label_escolaridade.grid(row=12, column=0, sticky="e", padx=(0, 10))
entry_escolaridade = tk.Entry(form_frame, textvariable=escolaridade_var, width=100)
entry_escolaridade.grid(row=12, column=1)

header_ministeriais = tk.Label(form_frame, text="Dados Ministeriais ", font=("Arial", 12, "bold"))
header_ministeriais.grid(row=13, columnspan=2, pady=(20, 5))

# Campos Dados Adicionais
label_cargo_na_igreja = tk.Label(form_frame, text="Exerce algum Cargo na Igreja/Qual?:")
label_cargo_na_igreja.grid(row=14, column=0, sticky="e", padx=(0, 10))
entry_cargo_na_igreja = tk.Entry(form_frame, textvariable=cargo_na_igreja_var, width=100)
entry_cargo_na_igreja.grid(row=14, column=1)

# Variáveis para armazenar os valores dos campos
data_batismo_var = tk.StringVar()
data_conversao_var = tk.StringVar()
data_batismo_data_var = tk.StringVar()
data_conversao_data_var = tk.StringVar()

# Labels e botões de rádio para o batismo nas águas
label_data_batismo = tk.Label(form_frame, text="Batizado nas águas?:")
label_data_batismo.grid(row=15, column=0, sticky="e", padx=(0, 10))

radio_data_batismo_sim = tk.Radiobutton(form_frame, text="Sim", variable=data_batismo_var, value="Sim")
radio_data_batismo_sim.grid(row=15, column=1, sticky="w")

radio_data_batismo_nao = tk.Radiobutton(form_frame, text="Não", variable=data_batismo_var, value="Não")
radio_data_batismo_nao.grid(row=16, column=1, sticky="w")

label_data_batismo_data = tk.Label(form_frame, text="Data do batismo nas águas?:")
label_data_batismo_data.grid(row=17, column=0, sticky="e", padx=(0, 10))
entry_data_batismo_data = tk.Entry(form_frame, textvariable=data_batismo_data_var, width=100)
entry_data_batismo_data.grid(row=17, column=1)

# Labels e botões de rádio para o batismo no Espírito Santo
label_data_conversao = tk.Label(form_frame, text="Batizado no Espírito Santo?:")
label_data_conversao.grid(row=18, column=0, sticky="e", padx=(0, 10))

data_conversao_sim = tk.Radiobutton(form_frame, text="Sim", variable=data_conversao_var, value="Sim")
data_conversao_sim.grid(row=18, column=1, sticky="w")

data_conversao_nao = tk.Radiobutton(form_frame, text="Não", variable=data_conversao_var, value="Não")
data_conversao_nao.grid(row=19, column=1, sticky="w")

label_data_conversao_data = tk.Label(form_frame, text="Data do batismo no Espírito Santo?:")
label_data_conversao_data.grid(row=20, column=0, sticky="e", padx=(0, 10))

entry_data_conversao_data = tk.Entry(form_frame, textvariable=data_conversao_data_var, width=100)
entry_data_conversao_data.grid(row=20, column=1)

label_data_ingresso = tk.Label(form_frame, text="Data de quando chegou em nossa igreja:")
label_data_ingresso.grid(row=21, column=0, sticky="w", padx=(10, 10))
entry_data_ingresso = tk.Entry(form_frame, textvariable=data_ingresso_var, width=50)
entry_data_ingresso.grid(row=21, column=1)


# Criar uma instância da classe GeradorExcel
gerador_excel = GeradorExcel("example.xlsx")

# Definir o diretório onde deseja salvar o arquivo Excel
gerador_excel.definir_diretorio(r"C:\Users\SAMSUNG\OneDrive\vs-python\cadig.py\imagens")

# Criar uma instância da classe GeradorGrafico
gerador_grafico = GeradorGrafico([])

# Botão para exibir os gráficos
button_exibir_graficos = ttk.Button(app, text="Exibir Gráficos", command=exibir_graficos)
button_exibir_graficos.grid(row=17, column=0, columnspan=2, pady=10)

# Botão Cadastrar
button_cadastrar = ttk.Button(app, text="Cadastrar", command=cadastrar_novo_membro)
button_cadastrar.grid(row=22, column=0, columnspan=2, pady=10)

# Executar o loop principal do Tkinter
app.mainloop()
